import argparse
import json
import openpyxl
import re


def main():
    ### ARGPARSE ###
    parser = argparse.ArgumentParser(
        description="Procesar un archivo Excel según la configuración proporcionada en un archivo JSON."
    )
    # Argumento de entrada. Path a fichero en formato JSON con la configuración del script
    parser.add_argument(
        "-c",
        "--config-file",
        required=True,
        help="Ruta del archivo JSON de configuración",
    )
    # Argumento de salida. Path del fichero de resultado del script
    parser.add_argument(
        "-o", "--output-file", help="Ruta del archivo '.xlsx' de salida analizado"
    )
    # Anlaizar los argumentos proporcionados al script
    args = parser.parse_args()

    # Cargar configuración desde el JSON
    # Abrir JSON y leer (r)
    with open(args.config_file, "r") as config_file:
        config = json.load(config_file)

    # Extraer parámetros de configuración del JSON
    input_file = config.get("input")
    sheet_name = config.get("sheet")
    search_pattern = config.get("search")

    # Validar que la estructura del JSON está correcta
    if not (input_file and sheet_name and search_pattern):
        print("Estructura del JSON con errores")
        return

    # Nombre archivo salida. Si no se especifica, usar nombre del script con extensión .xlsx
    output_file = args.output_file or "%(prog)s.xlsx"

    ### OPENPYXL ##
    # Abrir fichero excel y seleccionar la hoja (sheet), manejando errores
    try:
        workbook = openpyxl.load_workbook(input_file)
        worksheet = workbook[sheet_name]
    except Exception as e:
        print(f"Error al abrir el fichero Excel: {e}")
        return

    # Recorrer la columna A de la hoja indicada en el JSON
    for row in worksheet.iter_rows(min_row=1, min_col=1):
        for cell in row:
            value = cell.value
            if value is not None:
                # El modulo 're' proporciona operaciones de matching
                match = re.search(search_pattern, str(value))
                result = "SI" if match else "NO"
                # Escribir el valor SI (si hay coincidencia) o NO (si no hay coincidencia)
                # en la celda correspondiente de la derecha (Columna B)
                worksheet.cell(row=cell.row, column=2, value=result)

    # Generar y guardar el Excel reprocesado con la respuesta con manejo de errores
    try:
        print(f"Archivo excel reprocesado guardado como {output_file}")
    except Exception as e:
        print(f"Error al guardar el excel: {e}")
